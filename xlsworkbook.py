from openpyxl import load_workbook
from xlrd import open_workbook
from xlrd.sheet import ctype_text
from io import BytesIO

class ExcelWorkBook :

  def __init__ (self,filename='',fileobject=None):

    self.filename= filename
    self.fileobject= fileobject
    self.wb= None
    self.ws= None
    self.header= []
    self.data= []

    try:
      self.format= self.filename.split('.')[-1].lower()
    except:
      raise TypeError("%s's format is not recognized" % self.filename)

    if self.format == 'xls':
      try:
        self.wb= open_workbook(self.filename)
      except:
        self.wb= open_workbook(file_contents=self.fileobject.read())
      self.ws= self.wb.sheet_by_index( 0 )

    elif self.format == 'xlsx':
      try:
        self.wb= load_workbook(filename=self.filename,read_only=True,data_only=True)
      except:
        self.wb= load_workbook(filename=BytesIO(self.fileobject.read()),data_only=True)
      self.ws= self.wb.active

    else:
      raise TypeError('%s format is not supported' % self.format)

  def headers(self):
    return self.header
    
  def nsheets(self):
    if self.format == 'xls':
      return self.wb.nsheets
    elif self.format == 'xlsx':
      return len ( self.wb.get_sheet_names() )

  def nrows(self):
    if self.format == 'xls':
      return self.ws.nrows
    elif self.format == 'xlsx':
      return self.ws.get_highest_row()

  def ncols(self):
    if self.format == 'xls':
      return self.ws.ncols
    elif self.format == 'xlsx':
      return self.ws.get_highest_column()

  def sheet_names(self):
    if self.format == 'xls':
      return self.wb.sheet_names()
    elif self.format == 'xlsx':
      return self.wb.get_sheet_names()

  def sheet_by_index(self, index):
    if self.format == 'xls':
      self.ws= self.wb.sheet_by_index( index )
    elif self.format == 'xlsx':
      try:
        sheets= self.wb.get_sheet_names()
        self.ws= self.wb[sheets[index]]
      except:
        raise ValueError('sheet_by_index: index out of range')

  def sheet_by_name(self, name):
    if self.format == 'xls':
      self.ws= self.wb.sheet_by_name( name )
    elif self.format == 'xlsx':
      self.ws= self.wb.get_sheet_by_name ( name )

  def sheet_name(self):
    if self.format == 'xls':
      return self.ws.name
    elif self.format == 'xlsx':
      return self.ws.title

  def parse(self):
    if self.format == 'xls':
      for r in range(self.ws.nrows):
        c=0
        obj={}
        for cell in self.ws.row(r):
          if r == 0: 
            self.header.append (cell.value) 
          else:
            key= self.header[c]
            if cell.value == '':
              obj[key]= None
            else:
              obj[key]= cell.value
            #type_str= ctype_text.get(cell.ctype,'Unknown type')
          c+=1
        if r > 0: self.data.append (obj)

    elif self.format == 'xlsx':
      r=0
      for row in self.ws.rows:
        c=0
        obj={}
        for cell in row:
          if r == 0: 
            self.header.append (cell.value) 
          else:
            key= self.header[c]
            obj[key]= cell.value
          c+=1
        if r > 0: self.data.append (obj)
        r+=1

    return self.data
